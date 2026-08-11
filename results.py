"""Cálculo dos pesos, gravação dos resultados e montagem do dashboard.

Cobre as Etapas 3 e 4 do plano de ação:
  * 3.1 - vincula as respostas do formulário ao arquivo de resultados;
  * 3.2 - monta uma tabela de cálculo (simulando SE + PROCV) por auditoria;
  * 3.3 - calcula a nota final de cada auditoria;
  * 4.1 - mantém a aba 'Resultado Geral' em formato de banco de dados;
  * 4.2 - gera painéis (tabelas dinâmicas) por Empresa, Localidade e Área.
"""
from __future__ import annotations

import shutil
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from . import config
from .data_loader import Secao, todas_perguntas

# ---------------------------------------------------------------------------
# Estilos reutilizáveis
# ---------------------------------------------------------------------------
_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")
_TOTAL_FILL = PatternFill("solid", fgColor="FCE4D6")
_BOLD = Font(bold=True)
_WRAP = Alignment(wrap_text=True, vertical="top")

SHEET_RESULTADO = "Resultado Geral"
SHEET_PESOS = "Pesos"


@dataclass
class Auditoria:
    """Dados completos de uma auditoria preenchida no formulário."""

    responsavel: str
    data: str
    inspetor: str
    area: str
    centro: str
    empresa: str
    localidade: str
    tipo_centro: str
    respostas: dict[str, str]  # codigo -> SIM/NÃO/N/A
    evidencias: dict[str, str] = field(default_factory=dict)  # codigo -> caminho PNG
    evidencia_detalhes: dict[str, str] = field(default_factory=dict)  # codigo -> texto adicional
    registrado_em: str = ""

    def novo_id(self) -> str:
        return datetime.now().strftime("%Y%m%d%H%M%S%f")[:-3]


# ---------------------------------------------------------------------------
# Cálculo de pontuação (Etapa 3)
# ---------------------------------------------------------------------------


def pontuar(secoes: list[Secao], respostas: dict[str, str]) -> dict:
    """Aplica a regra: SIM -> peso da pergunta; NÃO/N/A -> 0.

    Devolve um resumo com pontuação por seção, nota final e contagens.
    """
    por_secao: dict[str, dict] = {}
    nota_final = 0.0
    peso_total = 0.0
    qtd_sim = qtd_nao = qtd_na = 0

    for s in secoes:
        pts = 0.0
        meta = 0.0
        for p in s.perguntas:
            meta += p.peso
            peso_total += p.peso
            resp = respostas.get(p.codigo, "")
            if resp == config.RESP_SIM:
                pts += p.peso
                qtd_sim += 1
            elif resp == config.RESP_NAO:
                qtd_nao += 1
            elif resp == config.RESP_NA:
                qtd_na += 1
        por_secao[s.titulo] = {"pontos": pts, "meta": meta}
        nota_final += pts

    aproveitamento = (nota_final / peso_total * 100.0) if peso_total else 0.0
    return {
        "por_secao": por_secao,
        "nota_final": nota_final,
        "peso_total": peso_total,
        "aproveitamento": aproveitamento,
        "qtd_sim": qtd_sim,
        "qtd_nao": qtd_nao,
        "qtd_na": qtd_na,
    }


# ---------------------------------------------------------------------------
# Gravação (Etapas 3 e 4)
# ---------------------------------------------------------------------------


def _copiar_evidencias(aud_id: str, evidencias: dict[str, str]) -> dict[str, str]:
    """Copia os PNGs de justificativa para uma pasta própria da auditoria."""
    if not evidencias:
        return {}
    destino = config.EVIDENCIAS_DIR / aud_id
    destino.mkdir(parents=True, exist_ok=True)
    salvos: dict[str, str] = {}
    for codigo, caminho in evidencias.items():
        src = Path(caminho)
        if not src.exists():
            continue
        alvo = destino / f"{codigo.replace('.', '_')}{src.suffix.lower()}"
        shutil.copy2(src, alvo)
        salvos[codigo] = str(alvo)
    return salvos


def _carregar_ou_criar_wb(secoes: list[Secao]) -> openpyxl.Workbook:
    if config.RESULTADOS_XLSX.exists():
        return openpyxl.load_workbook(config.RESULTADOS_XLSX)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _criar_aba_resultado(wb, secoes)
    _criar_aba_pesos(wb, secoes)
    return wb


def _colunas_resultado(secoes: list[Secao]) -> list[str]:
    base = [
        "Registrado em", "Área", "Centro", "Empresa",
        "Localidade", "Tipo de Centro", "Responsável", "Data", "Inspetor",
    ]
    secs = [f"Nota - {s.titulo}" for s in secoes]
    fim = ["Nota Final", "Peso Total", "Aproveitamento (%)", "Qtd SIM", "Qtd NÃO", "Qtd N/A"]
    return base + secs + fim


def _criar_aba_resultado(wb: openpyxl.Workbook, secoes: list[Secao]) -> None:
    ws = wb.create_sheet(SHEET_RESULTADO, 0)
    cols = _colunas_resultado(secoes)
    ws.append(cols)
    for c in range(1, len(cols) + 1):
        cell = ws.cell(1, c)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ws.freeze_panes = "A2"


def _criar_aba_pesos(wb: openpyxl.Workbook, secoes: list[Secao]) -> None:
    """Tabela de referência usada pelas fórmulas PROCV (VLOOKUP)."""
    ws = wb.create_sheet(SHEET_PESOS)
    ws.append(["Código", "Peso", "Seção", "Pergunta"])
    for c in range(1, 5):
        ws.cell(1, c).font = _BOLD
    for p in todas_perguntas(secoes):
        ws.append([p.codigo, p.peso, p.secao_titulo, p.texto])
    ws.sheet_state = "hidden"


def _sanitize_sheet_name(nome: str) -> str:
    """Remove caracteres inválidos e limita o nome a 31 caracteres para abas do Excel."""
    invalid_chars = set(r"\/:?*[]")
    nome_limpo = "".join(c for c in nome if c not in invalid_chars)
    nome_limpo = nome_limpo.strip()
    if not nome_limpo:
        nome_limpo = "Auditoria"
    return nome_limpo[:31]


def _unique_sheet_name(wb: openpyxl.Workbook, nome: str) -> str:
    base = nome
    contador = 1
    while nome in wb.sheetnames:
        sufixo = f"_{contador}"
        nome = base[:31 - len(sufixo)] + sufixo
        contador += 1
    return nome

def _send_email_via_outlook(caminho_relatorio: Path, destinatario: str, aud: Auditoria, nome_aba: str) -> None:
    try:
        import win32com.client as win32
    except Exception as e:  # pragma: no cover - environment dependent
        raise RuntimeError("PyWin32 (win32com) não está disponível: instale 'pywin32' para usar Outlook.") from e

    try:
        outlook = win32.Dispatch("Outlook.Application")
        mail = outlook.CreateItem(0)  # olMailItem
        mail.To = destinatario
        mail.Subject = f"Relatório final do inventário — {aud.centro} {aud.data}"
        body = (
            f"Olá,\n\n"
            f"Segue em anexo o relatório final do inventário.\n\n"
            f"Centro: {aud.centro}\n"
            f"Localidade: {aud.localidade}\n"
            f"Data: {aud.data}\n"
            f"Aba detalhada: {nome_aba}\n\n"
            f"Atenciosamente,\n"
            f"Equipe Inventário"
        )
        mail.Body = body
        mail.Attachments.Add(str(caminho_relatorio))
        mail.Send()
    except Exception as e:  # pragma: no cover - environment dependent
        raise RuntimeError(f"Falha ao enviar e-mail via Outlook: {e}") from e


def _send_email_relatorio(caminho_relatorio: Path, destinatario: str, aud: Auditoria, nome_aba: str) -> None:
    """Envia o relatório via Outlook/COM usando o cliente do Outlook local."""
    return _send_email_via_outlook(caminho_relatorio, destinatario, aud, nome_aba)


def _escrever_detalhe(wb, aud_id, aud, secoes, evid_salvas, nome: str):
    """Cria a aba de cálculo por auditoria (Ação 3.2) com fórmulas SE + PROCV."""
    ws = wb.create_sheet(nome)
    titulo = [
        ("Empresa", aud.empresa), ("Centro", aud.centro), ("Localidade", aud.localidade),
        ("Tipo de Centro", aud.tipo_centro), ("Área", aud.area),
        ("Responsável", aud.responsavel), ("Data", aud.data), ("Inspetor", aud.inspetor),
    ]
    for i, (rot, val) in enumerate(titulo, start=1):
        ws.cell(i, 1, rot).font = _BOLD
        ws.cell(i, 2, val)
    header_row = len(titulo) + 2
    cabec = ["Código", "Seção", "Pergunta", "Peso", "Resposta", "Pontuado", "Evidência", "Justificativa", "Plano de ação"]
    for c, txt in enumerate(cabec, start=1):
        cell = ws.cell(header_row, c, txt)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
    r = header_row + 1
    primeira = r
    for s in secoes:
        for p in s.perguntas:
            resp = aud.respostas.get(p.codigo, "")
            ws.cell(r, 1, p.codigo)
            ws.cell(r, 2, p.secao_titulo)
            ws.cell(r, 3, p.texto).alignment = _WRAP
            ws.cell(r, 4, p.peso)
            ws.cell(r, 5, resp)
            # Ação 3.2: SE(resposta="SIM"; PROCV(codigo; Pesos; 2; 0); 0)
            ws.cell(r, 6).value = (
                f'=IF(E{r}="{config.RESP_SIM}",'
                f'VLOOKUP(A{r},{SHEET_PESOS}!$A:$B,2,FALSE),0)'
            )
            evid = evid_salvas.get(p.codigo, "")
            ws.cell(r, 7, Path(evid).name if evid else "")
            detalhe = aud.evidencia_detalhes.get(p.codigo, "")
            ws.cell(r, 8, detalhe).alignment = _WRAP
            if resp == config.RESP_NAO and p.acao:
                ws.cell(r, 9, p.acao).alignment = _WRAP
            r += 1
    ultima = r - 1
    # Ação 3.3: célula de somatória final.
    ws.cell(r, 3, "NOTA FINAL").font = _BOLD
    total_cell = ws.cell(r, 6)
    total_cell.value = f"=SUM(F{primeira}:F{ultima})"
    total_cell.font = _BOLD
    total_cell.fill = _TOTAL_FILL
    _ajustar_larguras(ws, {1: 10, 2: 26, 3: 60, 4: 8, 5: 12, 6: 12, 7: 22, 8: 40, 9: 40})
    return nome


def _append_resultado(wb, aud_id, aud, secoes, resumo):
    ws = wb[SHEET_RESULTADO]
    linha = [
        aud.registrado_em, aud.area, aud.centro, aud.empresa,
        aud.localidade, aud.tipo_centro, aud.responsavel, aud.data, aud.inspetor,
    ]
    for s in secoes:
        linha.append(round(resumo["por_secao"][s.titulo]["pontos"], 3))
    linha += [
        round(resumo["nota_final"], 3),
        round(resumo["peso_total"], 3),
        round(resumo["aproveitamento"], 1),
        resumo["qtd_sim"], resumo["qtd_nao"], resumo["qtd_na"],
    ]
    ws.append(linha)


def _ajustar_larguras(ws, larguras: dict[int, int]):
    for col, w in larguras.items():
        ws.column_dimensions[get_column_letter(col)].width = w


# ---------------------------------------------------------------------------
# Ponto de entrada da gravação
# ---------------------------------------------------------------------------


def salvar_auditoria(aud: Auditoria, secoes: list[Secao]) -> dict:
    """Grava a auditoria (linha no banco de dados + aba de cálculo)."""
    config.ensure_output_dirs()
    aud_id = aud.novo_id()
    aud.registrado_em = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    resumo = pontuar(secoes, aud.respostas)

    wb = _carregar_ou_criar_wb(secoes)
    if SHEET_RESULTADO not in wb.sheetnames:
        _criar_aba_resultado(wb, secoes)
    if SHEET_PESOS not in wb.sheetnames:
        _criar_aba_pesos(wb, secoes)

    nome_detalhe = _sanitize_sheet_name(f"{aud.centro} {aud.data}")
    nome_detalhe = _unique_sheet_name(wb, nome_detalhe)
    evid_salvas = _copiar_evidencias(config.EVIDENCIAS_DIR / nome_detalhe, aud.evidencias)

    _append_resultado(wb, aud_id, aud, secoes, resumo)
    nome_detalhe = _escrever_detalhe(wb, aud_id, aud, secoes, evid_salvas, nome_detalhe)

    config.RESULTADOS_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(config.RESULTADOS_XLSX)

    return {
        "id": aud_id,
        "arquivo": str(config.RESULTADOS_XLSX),
        "aba_detalhe": nome_detalhe,
        "resumo": resumo,
    }
