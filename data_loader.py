"""Leitura das bases de dados (Cronograma e Checklist).

Responsável por:
  * Ação 1.2 - montar a lista única de combinações válidas de
    Área / Centro / Empresa / Localidade / Tipo de Centro.
  * Ação 2.2 - alimentar as listas suspensas do formulário.
  * Ação 2.3 / 2.5 - carregar as perguntas do checklist com seus pesos.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field

import openpyxl

from . import config


# ---------------------------------------------------------------------------
# Combinações válidas (Cronograma Atualizado)
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class Combinacao:
    """Uma combinação válida vinda do Cronograma."""

    area: str
    centro: str
    empresa: str
    localidade: str
    tipo_centro: str

    def as_dict(self) -> dict[str, str]:
        return {
            "area": self.area,
            "centro": self.centro,
            "empresa": self.empresa,
            "localidade": self.localidade,
            "tipo_centro": self.tipo_centro,
        }


def _clean(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def carregar_combinacoes(path=None) -> list[Combinacao]:
    """Lê o Cronograma e devolve a lista única de combinações válidas (Ação 1.2)."""
    path = path or config.CRONOGRAMA_PATH
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[config.CRONOGRAMA_SHEET]

    col = {k: openpyxl.utils.column_index_from_string(v) for k, v in config.CRONOGRAMA_COLS.items()}

    vistos: set[Combinacao] = set()
    combinacoes: list[Combinacao] = []
    for r in range(config.CRONOGRAMA_HEADER_ROW + 1, ws.max_row + 1):
        area = _clean(ws.cell(r, col["area"]).value)
        centro = _clean(ws.cell(r, col["centro"]).value)
        empresa = _clean(ws.cell(r, col["empresa"]).value)
        localidade = _clean(ws.cell(r, col["localidade"]).value)
        tipo = _clean(ws.cell(r, col["tipo_centro"]).value)
        # Linha precisa ter ao menos os campos-chave preenchidos.
        if not (area and centro and empresa):
            continue
        combo = Combinacao(area, centro, empresa, localidade, tipo)
        if combo not in vistos:
            vistos.add(combo)
            combinacoes.append(combo)
    wb.close()
    return combinacoes


class ReferenciaCombinacoes:
    """Facilita as consultas em cascata das listas suspensas (Ação 2.2)."""

    def __init__(self, combinacoes: list[Combinacao]):
        self.combinacoes = combinacoes

    def _filtrar(self, selecao: dict[str, str]) -> list[Combinacao]:
        resultado = []
        for c in self.combinacoes:
            d = c.as_dict()
            if all(not v or d.get(k) == v for k, v in selecao.items()):
                resultado.append(c)
        return resultado

    def opcoes(self, campo: str, selecao: dict[str, str]) -> list[str]:
        """Valores possíveis para ``campo`` dadas as seleções anteriores."""
        filtradas = self._filtrar(selecao)
        valores = {c.as_dict()[campo] for c in filtradas if c.as_dict()[campo]}
        return sorted(valores)

    def combinacao_valida(self, selecao: dict[str, str]) -> bool:
        return len(self._filtrar(selecao)) > 0


# ---------------------------------------------------------------------------
# Perguntas do checklist (Questionário)
# ---------------------------------------------------------------------------


@dataclass
class Pergunta:
    codigo: str  # ex.: "1.1"
    texto: str
    peso: float
    secao_num: int
    secao_titulo: str
    acao: str = ""  # texto do plano de ação sugerido (coluna N)


@dataclass
class Secao:
    numero: int
    titulo: str
    meta: float
    perguntas: list[Pergunta] = field(default_factory=list)


_CODIGO_RE = re.compile(r"^\d+\.\d+$")


def carregar_questionario(path=None) -> list[Secao]:
    """Lê a aba 'Questionário' e devolve as seções com suas perguntas e pesos."""
    path = path or config.CHECKLIST_PATH
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[config.CHECKLIST_SHEET]

    # Colunas conhecidas da planilha original.
    C_SECAO_NUM = 2   # B
    C_CODIGO = 3      # C (código da pergunta OU título da seção)
    C_TEXTO = 4       # D
    C_PESO = 10       # J
    C_META = 12       # L
    C_ACAO = 14       # N

    secoes: list[Secao] = []
    atual: Secao | None = None

    for r in range(1, ws.max_row + 1):
        b = ws.cell(r, C_SECAO_NUM).value
        c = _clean(ws.cell(r, C_CODIGO).value)
        d = _clean(ws.cell(r, C_TEXTO).value)

        # Linha de cabeçalho de seção: B é um número inteiro e C traz o título.
        if isinstance(b, (int, float)) and c and not _CODIGO_RE.match(c):
            meta = ws.cell(r, C_META).value
            atual = Secao(numero=int(b), titulo=c, meta=float(meta) if meta else 0.0)
            secoes.append(atual)
            continue

        # Linha de pergunta: C no formato "x.y".
        if _CODIGO_RE.match(c) and atual is not None:
            peso = ws.cell(r, C_PESO).value
            acao = _clean(ws.cell(r, C_ACAO).value)
            atual.perguntas.append(
                Pergunta(
                    codigo=c,
                    texto=d,
                    peso=float(peso) if peso else 0.0,
                    secao_num=atual.numero,
                    secao_titulo=atual.titulo,
                    acao=acao,
                )
            )
    wb.close()
    return secoes


def todas_perguntas(secoes: list[Secao]) -> list[Pergunta]:
    return [p for s in secoes for p in s.perguntas]


def peso_total(secoes: list[Secao]) -> float:
    return sum(p.peso for p in todas_perguntas(secoes))
